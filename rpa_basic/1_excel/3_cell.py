from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1이라는 값을 입력함.
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1셀의 정보를 출력함
print(ws["A1"].value) # A1셀의 값을 출력함.
print(ws["A10"].value) # 값을 넣지 않았을 시 터미널에 None을 출력함.

# row = 1, 2, 3, ....
# column =  A, B, C, ....
print(ws.cell(row=1, column=1).value) # 첫번째 줄의 1번 컬럼이므로 A1을 의미함. ws["A1"].value임.

print(ws.cell(row=1, column=2).value) # ws["B1"].value


c = ws.cell(column=3, row=1, value = 10) # ws["C1"].value = 10 

print(c.value) # ws["C1"]

from random import * # 원래는 위에 넣어야 할 코드이나 강의 진행상 여기다 넣음.

# 이중 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10개 column
        #ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")