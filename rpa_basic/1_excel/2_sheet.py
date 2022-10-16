from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "Mysheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff00ff" # RGB형태로 값을 넣어주면 탭 색상 변경됨.

# Sheet, Mysheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] # Dict형태로 Sheet에 접근하기

print(wb.sheetnames) # 모든 sheet이름 확인하는 출력기능

# Sheet 복사
new_ws["A3"] = "Test" #액셀 문서상의 맨 위 A1셀의 이름을 Test로 바꿈.
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

# 위 코드 설명 : new_ws를 복사해서 그 복사된 시트를 target이라는 워크시트로 저장(변수에 담아서)하고, 제목을 Copied Sheet로 바꿔서 저장해준다. 


wb.save("sample.xlsx")